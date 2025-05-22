#!/usr/bin/env python3
"""
xlsx-translator - Excel Translation Utility with Gemini API

A command-line tool for translating Microsoft Excel files while preserving formatting,
formulas, styles, and structure. Supports batch processing, concurrent API calls,
and various filtering options.

Features:
- Translate single files or entire directories
- Preserve cell formatting, formulas, merged cells, and images
- Support for both .xlsx and .xlsm (macro-enabled) files
- Filter by sheet, rows, or columns
- Context-aware translation with glossary support
- Multiple style prompts for different translation tones
- Concurrent API calls for faster processing

Usage Examples:
    # Translate a single file to Japanese
    xlsx-translator input.xlsx --target-lang ja

    # Translate directory with multiple target languages
    xlsx-translator --source-location ./excel_files --target-langs en,ja,es

    # Translate specific sheet and columns with business style
    xlsx-translator report.xlsx --sheet-name "Sales" --columns "A,C:E" --target-lang en --style-prompt business

    # Use context file for consistent terminology
    xlsx-translator technical.xlsx --target-lang ja --context-file glossary.json

    # List available languages and styles
    xlsx-translator --list-languages
    xlsx-translator --list-styles
"""

import argparse
import asyncio
import json
import logging
import os
import re
import shutil
import sys
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from copy import copy
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Optional, Set, Tuple, Union

import google.generativeai as genai
import openpyxl
from colorama import Fore, Style, init
from openpyxl.utils import column_index_from_string, get_column_letter
from tqdm import tqdm

# Initialize colorama for colored output
init(autoreset=True)

# Constants
DEFAULT_BATCH_SIZE = 20
DEFAULT_CONCURRENCY = 4
MAX_RETRIES = 3
INVALID_SHEET_CHARS = r'[:\\/?*\[\]]'
MAX_SHEET_NAME_LENGTH = 31

# Supported languages (subset - extend as needed)
SUPPORTED_LANGUAGES = {
    'en': 'English',
    'es': 'Spanish',
    'fr': 'French',
    'de': 'German',
    'it': 'Italian',
    'pt': 'Portuguese',
    'ru': 'Russian',
    'ja': 'Japanese',
    'ko': 'Korean',
    'zh': 'Chinese (Simplified)',
    'zh-TW': 'Chinese (Traditional)',
    'ar': 'Arabic',
    'hi': 'Hindi',
    'tr': 'Turkish',
    'pl': 'Polish',
    'nl': 'Dutch',
    'sv': 'Swedish',
    'da': 'Danish',
    'no': 'Norwegian',
    'fi': 'Finnish',
}

# Style prompts
STYLE_PROMPTS = {
    'business': 'Translate using formal business language appropriate for professional documents.',
    'casual': 'Translate using conversational, friendly language suitable for informal communication.',
    'technical': 'Translate using precise technical terminology, maintaining accuracy for specialized content.',
    'marketing': 'Translate using persuasive, engaging language suitable for marketing materials.',
}


@dataclass
class TranslationConfig:
    """Configuration for translation job"""
    source_lang: str
    target_lang: str
    model_name: str
    batch_size: int
    concurrency: int
    preserve_formatting: bool
    preserve_formulas: bool
    style_prompt: str
    context_file: Optional[str]
    max_context_items: int
    debug: bool


@dataclass
class CellFilter:
    """Filtering configuration for cells"""
    sheet_name: Optional[str]
    sheet_index: Optional[int]
    columns: Optional[Set[int]]
    rows: Optional[Set[int]]


class ExcelTranslator:
    """Main translator class"""
    
    def __init__(self, config: TranslationConfig):
        self.config = config
        self.context_data = self._load_context()
        self._setup_logging()
        self._configure_api()
        
    def _setup_logging(self):
        """Configure logging based on debug flag"""
        level = logging.DEBUG if self.config.debug else logging.INFO
        logging.basicConfig(
            level=level,
            format='%(asctime)s - %(levelname)s - %(message)s'
        )
        self.logger = logging.getLogger(__name__)
        
    def _configure_api(self):
        """Configure Gemini API"""
        api_key = os.getenv('GOOGLE_API_KEY')
        if not api_key:
            raise ValueError(
                f"{Fore.RED}Error: GOOGLE_API_KEY environment variable not set.{Style.RESET_ALL}\n"
                f"{Fore.YELLOW}Please set your Gemini API key as an environment variable.{Style.RESET_ALL}"
            )
        genai.configure(api_key=api_key)
        self.model = genai.GenerativeModel(self.config.model_name)
        
    def _load_context(self) -> Dict:
        """Load context/glossary file if provided"""
        if not self.config.context_file:
            return {}
            
        try:
            with open(self.config.context_file, 'r', encoding='utf-8') as f:
                data = json.load(f)
                self.logger.info(f"Loaded context from {self.config.context_file}")
                return data
        except Exception as e:
            self.logger.warning(f"Failed to load context file: {e}")
            return {}
    
    def _build_prompt(self, texts: List[Tuple[int, str]]) -> str:
        """Build translation prompt with context and style"""
        batch_data = [{"id": i, "text": text} for i, text in texts]
        
        context_str = ""
        if self.context_data:
            # Include relevant context items
            context_items = list(self.context_data.items())[:self.config.max_context_items]
            if context_items:
                context_str = "\nGlossary/Context:\n"
                for term, translation in context_items:
                    context_str += f"- {term} → {translation}\n"
        
        style_instruction = STYLE_PROMPTS.get(self.config.style_prompt, "")
        
        prompt = f"""
Translate from {self.config.source_lang} to {self.config.target_lang}.
{style_instruction}
{context_str}
IMPORTANT:
- Maintain ALL formatting EXACTLY (spaces, punctuation, capitalization patterns)
- For empty strings, return empty strings
- For numbers only, return the same numbers
- Apply glossary terms consistently

Return ONLY valid JSON in this exact format:
{{
    "translations": [
        {{"id": <id>, "translation": "<translated_text>"}}
    ]
}}

Input data:
{json.dumps(batch_data, ensure_ascii=False)}
"""
        return prompt
    
    def _extract_json(self, response_text: str) -> str:
        """Extract JSON from response"""
        # Try to find JSON code blocks
        matches = re.findall(r'```(?:json)?\n(.*?)\n```', response_text, re.DOTALL)
        if matches:
            return matches[0]
        
        # Try to find raw JSON
        match = re.search(r'{.*}', response_text, re.DOTALL)
        if match:
            return match.group(0)
            
        return response_text
    
    def translate_batch(self, texts: List[Tuple[int, str]]) -> Dict[int, str]:
        """Translate a batch of texts with retry logic"""
        if not texts:
            return {}
            
        for attempt in range(MAX_RETRIES):
            try:
                prompt = self._build_prompt(texts)
                response = self.model.generate_content(prompt)
                
                if self.config.debug:
                    self.logger.debug(f"API Response: {response.text[:200]}...")
                
                cleaned = self._extract_json(response.text)
                result = json.loads(cleaned)
                
                if 'translations' not in result:
                    raise ValueError("Missing 'translations' key in response")
                
                # Build translation map
                translation_map = {}
                for item in result['translations']:
                    idx = item.get('id')
                    translation = item.get('translation', '')
                    if idx is not None:
                        translation_map[idx] = translation
                
                return translation_map
                
            except Exception as e:
                self.logger.warning(f"Translation attempt {attempt + 1} failed: {e}")
                if attempt < MAX_RETRIES - 1:
                    time.sleep(2 ** attempt)  # Exponential backoff
                else:
                    self.logger.error(f"All translation attempts failed for batch")
                    return {i: text for i, text in texts}  # Return originals
    
    def translate_file(self, input_path: Path, output_path: Path, cell_filter: CellFilter):
        """Translate a single Excel file"""
        self.logger.info(f"Processing: {input_path}")
        
        try:
            # Load workbook
            wb = openpyxl.load_workbook(input_path, data_only=False, keep_vba=True)
            
            # Create temporary translated workbook
            temp_file = output_path.with_suffix('.tmp' + output_path.suffix)
            wb_trans = openpyxl.Workbook()
            if 'Sheet' in wb_trans.sheetnames:
                del wb_trans['Sheet']
            
            # Get sheets to process
            sheets_to_process = self._get_sheets_to_process(wb, cell_filter)
            
            # Process each sheet
            for sheet_name in tqdm(sheets_to_process, desc="Sheets", leave=False):
                self._process_sheet(wb[sheet_name], wb_trans, cell_filter)
            
            # Save translated workbook
            wb_trans.save(temp_file)
            wb_trans.close()
            wb.close()
            
            # Apply formatting preservation
            if self.config.preserve_formatting:
                self._preserve_formatting(input_path, temp_file, output_path)
                temp_file.unlink()  # Remove temp file
            else:
                shutil.move(temp_file, output_path)
            
            self.logger.info(f"✅ Created: {output_path}")
            
        except Exception as e:
            self.logger.error(f"Failed to process {input_path}: {e}")
            raise
    
    def _get_sheets_to_process(self, wb: openpyxl.Workbook, cell_filter: CellFilter) -> List[str]:
        """Determine which sheets to process based on filter"""
        if cell_filter.sheet_name:
            if cell_filter.sheet_name in wb.sheetnames:
                return [cell_filter.sheet_name]
            else:
                raise ValueError(f"Sheet '{cell_filter.sheet_name}' not found")
        elif cell_filter.sheet_index is not None:
            if 0 <= cell_filter.sheet_index < len(wb.sheetnames):
                return [wb.sheetnames[cell_filter.sheet_index]]
            else:
                raise ValueError(f"Sheet index {cell_filter.sheet_index} out of range")
        else:
            return wb.sheetnames
    
    def _process_sheet(self, source_sheet, target_wb, cell_filter: CellFilter):
        """Process a single worksheet"""
        # Translate sheet name
        sheet_name_trans = self.translate_batch([(0, source_sheet.title)])
        new_name = sheet_name_trans.get(0, source_sheet.title)
        new_name = re.sub(INVALID_SHEET_CHARS, '_', new_name)[:MAX_SHEET_NAME_LENGTH]
        
        target_sheet = target_wb.create_sheet(new_name)
        
        # Copy sheet properties
        target_sheet.sheet_properties = copy(source_sheet.sheet_properties)
        
        # Collect all texts to translate
        texts_to_translate = []
        cell_mapping = {}
        
        for row in source_sheet.iter_rows():
            # Check row filter
            if cell_filter.rows and row[0].row not in cell_filter.rows:
                continue
                
            for cell in row:
                # Check column filter
                if cell_filter.columns and cell.column not in cell_filter.columns:
                    continue
                
                # Copy cell to target
                target_cell = target_sheet.cell(row=cell.row, column=cell.column)
                
                # Copy style if preserving formatting
                if self.config.preserve_formatting and cell.has_style:
                    target_cell.font = copy(cell.font)
                    target_cell.border = copy(cell.border)
                    target_cell.fill = copy(cell.fill)
                    target_cell.number_format = cell.number_format
                    target_cell.alignment = copy(cell.alignment)
                
                # Handle cell value
                if cell.value is None:
                    continue
                elif self.config.preserve_formulas and str(cell.value).startswith('='):
                    target_cell.value = cell.value  # Keep formula
                elif isinstance(cell.value, str) and cell.value.strip():
                    # Add to translation batch
                    batch_id = len(texts_to_translate)
                    texts_to_translate.append((batch_id, cell.value))
                    cell_mapping[batch_id] = (cell.row, cell.column)
                else:
                    target_cell.value = cell.value  # Copy as-is (numbers, dates, etc.)
        
        # Translate in batches with concurrency
        self._translate_cells_concurrent(texts_to_translate, cell_mapping, target_sheet)
        
        # Copy other sheet properties
        self._copy_sheet_properties(source_sheet, target_sheet)
    
    def _translate_cells_concurrent(self, texts: List[Tuple[int, str]], 
                                   cell_mapping: Dict[int, Tuple[int, int]], 
                                   target_sheet):
        """Translate cells using concurrent API calls"""
        if not texts:
            return
            
        # Split into batches
        batches = []
        for i in range(0, len(texts), self.config.batch_size):
            batches.append(texts[i:i + self.config.batch_size])
        
        # Process batches concurrently
        all_translations = {}
        with ThreadPoolExecutor(max_workers=self.config.concurrency) as executor:
            future_to_batch = {
                executor.submit(self.translate_batch, batch): batch 
                for batch in batches
            }
            
            for future in tqdm(as_completed(future_to_batch), 
                             total=len(batches), 
                             desc="Translating", 
                             leave=False):
                translations = future.result()
                all_translations.update(translations)
        
        # Apply translations to cells
        for batch_id, translation in all_translations.items():
            if batch_id in cell_mapping:
                row, col = cell_mapping[batch_id]
                target_sheet.cell(row=row, column=col).value = translation
    
    def _copy_sheet_properties(self, source_sheet, target_sheet):
        """Copy sheet properties like column widths, row heights, etc."""
        # Copy column dimensions
        for col, dim in source_sheet.column_dimensions.items():
            target_sheet.column_dimensions[col] = copy(dim)
        
        # Copy row dimensions
        for row, dim in source_sheet.row_dimensions.items():
            target_sheet.row_dimensions[row] = copy(dim)
        
        # Copy print settings
        target_sheet.print_options = copy(source_sheet.print_options)
        target_sheet.page_setup = copy(source_sheet.page_setup)
        target_sheet.page_margins = copy(source_sheet.page_margins)
    
    def _preserve_formatting(self, original_file: Path, translated_file: Path, output_file: Path):
        """Preserve formatting by copying translated values to original structure"""
        try:
            # Copy original to output
            shutil.copy2(original_file, output_file)
            time.sleep(0.1)  # Brief pause for file system
            
            # Open files
            wb_trans = openpyxl.load_workbook(translated_file, data_only=True)
            wb_final = openpyxl.load_workbook(output_file, keep_vba=True)
            
            # Copy translated content
            for idx, sheet_name in enumerate(wb_trans.sheetnames):
                if idx < len(wb_final.worksheets):
                    trans_sheet = wb_trans.worksheets[idx]
                    final_sheet = wb_final.worksheets[idx]
                    
                    # Update sheet name
                    try:
                        final_sheet.title = trans_sheet.title
                    except:
                        pass  # Keep original if rename fails
                    
                    # Handle merged cells
                    merged_ranges = list(final_sheet.merged_cells.ranges)
                    for merged_range in merged_ranges:
                        final_sheet.unmerge_cells(str(merged_range))
                    
                    # Copy values
                    for row in trans_sheet.iter_rows():
                        for cell in row:
                            if cell.value is not None:
                                final_sheet.cell(row=cell.row, column=cell.column).value = cell.value
                    
                    # Re-merge cells
                    for merged_range in merged_ranges:
                        final_sheet.merge_cells(str(merged_range))
            
            # Save final file
            wb_final.save(output_file)
            wb_final.close()
            wb_trans.close()
            
        except Exception as e:
            self.logger.error(f"Format preservation failed: {e}")
            # Fall back to translated file
            shutil.copy2(translated_file, output_file)


def parse_columns(columns_str: str) -> Set[int]:
    """Parse column specification like 'A,C:E,G' into column indices"""
    columns = set()
    for part in columns_str.split(','):
        part = part.strip()
        if ':' in part:
            # Range like C:E
            start, end = part.split(':')
            start_idx = column_index_from_string(start)
            end_idx = column_index_from_string(end)
            columns.update(range(start_idx, end_idx + 1))
        else:
            # Single column
            columns.add(column_index_from_string(part))
    return columns


def parse_rows(rows_str: str) -> Set[int]:
    """Parse row specification like '1,5-10,15' into row indices"""
    rows = set()
    for part in rows_str.split(','):
        part = part.strip()
        if '-' in part:
            # Range like 5-10
            start, end = map(int, part.split('-'))
            rows.update(range(start, end + 1))
        else:
            # Single row
            rows.add(int(part))
    return rows


def process_directory(translator: ExcelTranslator, source_dir: Path, output_dir: Path, 
                     cell_filter: CellFilter, target_langs: List[str]):
    """Process all Excel files in a directory"""
    excel_files = list(source_dir.glob('*.xlsx')) + list(source_dir.glob('*.xlsm'))
    
    if not excel_files:
        print(f"{Fore.YELLOW}No Excel files found in {source_dir}{Style.RESET_ALL}")
        return
    
    print(f"{Fore.GREEN}Found {len(excel_files)} Excel files{Style.RESET_ALL}")
    
    for target_lang in target_langs:
        print(f"\n{Fore.CYAN}Translating to {target_lang} ({SUPPORTED_LANGUAGES.get(target_lang, 'Unknown')}){Style.RESET_ALL}")
        
        # Update translator config
        translator.config.target_lang = target_lang
        
        # Create language subdirectory
        lang_dir = output_dir / target_lang
        lang_dir.mkdir(parents=True, exist_ok=True)
        
        # Process each file
        for excel_file in tqdm(excel_files, desc=f"Files ({target_lang})"):
            output_name = f"{excel_file.stem}_{target_lang}{excel_file.suffix}"
            output_path = lang_dir / output_name
            
            try:
                translator.translate_file(excel_file, output_path, cell_filter)
            except Exception as e:
                print(f"{Fore.RED}Failed to process {excel_file}: {e}{Style.RESET_ALL}")


def main():
    """Main entry point"""
    parser = argparse.ArgumentParser(
        description="Excel Translation Utility with Gemini API",
        epilog="Examples:\n"
               "  %(prog)s input.xlsx --target-lang ja\n"
               "  %(prog)s --source-location ./files --target-langs en,ja,es\n"
               "  %(prog)s report.xlsx --sheet-name Sales --columns A,C:E --target-lang en",
        formatter_class=argparse.RawDescriptionHelpFormatter
    )
    
    # Input options (mutually exclusive)
    input_group = parser.add_mutually_exclusive_group(required=False)
    input_group.add_argument('input_file', nargs='?', help='Input Excel file (.xlsx or .xlsm)')
    input_group.add_argument('--source-location', type=str, help='Directory containing Excel files')
    
    # Language options
    parser.add_argument('--source-lang', default='en', help='Source language code (default: en)')
    
    lang_group = parser.add_mutually_exclusive_group()
    lang_group.add_argument('--target-lang', help='Target language code (e.g., ja)')
    lang_group.add_argument('--target-langs', help='Multiple target languages (comma-separated, e.g., en,ja,es)')
    
    # Excel-specific options
    parser.add_argument('--sheet-name', help='Process only this sheet by name')
    parser.add_argument('--sheet-index', type=int, help='Process only this sheet by index (0-based)')
    parser.add_argument('--columns', help='Columns to translate (e.g., "A,C:E,G")')
    parser.add_argument('--rows', help='Rows to translate (e.g., "1,5-10,15")')
    parser.add_argument('--preserve-formatting', action='store_true', default=True,
                       help='Preserve cell formatting (default: True)')
    parser.add_argument('--preserve-formulas', action='store_true', default=True,
                       help='Skip cells containing formulas (default: True)')
    
    # Context and styling
    parser.add_argument('--context-file', help='Path to glossary/context JSON file')
    parser.add_argument('--style-prompt', choices=list(STYLE_PROMPTS.keys()), 
                       default='business', help='Translation style (default: business)')
    parser.add_argument('--max-context-items', type=int, default=10,
                       help='Maximum context items to include (default: 10)')
    
    # Output and processing
    parser.add_argument('--output-dir', default='.', help='Output directory (default: current)')
    parser.add_argument('--model', default='gemini-2.0-flash', help='Gemini model name')
    parser.add_argument('--batch-size', type=int, default=DEFAULT_BATCH_SIZE,
                       help=f'Cells per API call (default: {DEFAULT_BATCH_SIZE})')
    parser.add_argument('--concurrency', type=int, default=DEFAULT_CONCURRENCY,
                       help=f'Concurrent API calls (default: {DEFAULT_CONCURRENCY})')
    
    # Info flags
    parser.add_argument('--list-styles', action='store_true', help='List available style options')
    parser.add_argument('--list-languages', action='store_true', help='List supported languages')
    parser.add_argument('--debug', action='store_true', help='Enable debug logging')
    
    args = parser.parse_args()
    
    # Handle info flags
    if args.list_styles:
        print(f"{Fore.CYAN}Available style prompts:{Style.RESET_ALL}")
        for style, description in STYLE_PROMPTS.items():
            print(f"  {Fore.GREEN}{style:<12}{Style.RESET_ALL} - {description}")
        return
    
    if args.list_languages:
        print(f"{Fore.CYAN}Supported languages:{Style.RESET_ALL}")
        for code, name in sorted(SUPPORTED_LANGUAGES.items()):
            print(f"  {Fore.GREEN}{code:<8}{Style.RESET_ALL} - {name}")
        return
    
    # Validate input
    if not args.input_file and not args.source_location:
        parser.error("Either input_file or --source-location is required")
    
    if not args.target_lang and not args.target_langs:
        parser.error("Either --target-lang or --target-langs is required")
    
    # Parse target languages
    if args.target_langs:
        target_langs = [lang.strip() for lang in args.target_langs.split(',')]
    else:
        target_langs = [args.target_lang]
    
    # Validate languages
    for lang in target_langs:
        if lang not in SUPPORTED_LANGUAGES:
            print(f"{Fore.YELLOW}Warning: '{lang}' may not be fully supported{Style.RESET_ALL}")
    
    # Create configuration
    config = TranslationConfig(
        source_lang=args.source_lang,
        target_lang=target_langs[0],  # Will be updated for each language
        model_name=args.model,
        batch_size=args.batch_size,
        concurrency=args.concurrency,
        preserve_formatting=args.preserve_formatting,
        preserve_formulas=args.preserve_formulas,
        style_prompt=args.style_prompt,
        context_file=args.context_file,
        max_context_items=args.max_context_items,
        debug=args.debug
    )
    
    # Create cell filter
    cell_filter = CellFilter(
        sheet_name=args.sheet_name,
        sheet_index=args.sheet_index,
        columns=parse_columns(args.columns) if args.columns else None,
        rows=parse_rows(args.rows) if args.rows else None
    )
    
    # Create translator
    try:
        translator = ExcelTranslator(config)
    except ValueError as e:
        print(str(e))
        sys.exit(1)
    
    # Process files
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    
    try:
        if args.input_file:
            # Single file mode
            input_path = Path(args.input_file)
            if not input_path.exists():
                print(f"{Fore.RED}Error: File not found: {input_path}{Style.RESET_ALL}")
                sys.exit(1)
            
            for target_lang in target_langs:
                translator.config.target_lang = target_lang
                output_name = f"{input_path.stem}_{target_lang}{input_path.suffix}"
                output_path = output_dir / output_name
                
                print(f"{Fore.CYAN}Translating to {target_lang}...{Style.RESET_ALL}")
                translator.translate_file(input_path, output_path, cell_filter)
        
        else:
            # Directory mode
            source_dir = Path(args.source_location)
            if not source_dir.exists():
                print(f"{Fore.RED}Error: Directory not found: {source_dir}{Style.RESET_ALL}")
                sys.exit(1)
            
            process_directory(translator, source_dir, output_dir, cell_filter, target_langs)
        
        print(f"\n{Fore.GREEN}✅ Translation complete!{Style.RESET_ALL}")
        
    except Exception as e:
        print(f"\n{Fore.RED}❌ Translation failed: {e}{Style.RESET_ALL}")
        if args.debug:
            import traceback
            traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()
