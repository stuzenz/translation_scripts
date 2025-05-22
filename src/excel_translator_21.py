#!/usr/bin/env python3
"""
xlsx-translator - Excel Translation Utility with Gemini API

A command-line tool for translating Microsoft Excel files while preserving formatting,
formulas, styles, and structure. Supports batch processing, concurrent API calls,
and various filtering options.

Features:
- Translate single files or entire directories
- Preserve cell formatting, formulas, merged cells, and images
- Support for both .xlsx and .xlsm (macro-enabled) files
- Filter by sheet, rows, or columns
- Context-aware translation with glossary support
- Multiple style prompts for different translation tones
- Concurrent API calls for faster processing
- Automatic context extraction from headers and surrounding cells

Usage Examples:
    # Translate a single file to Japanese
    xlsx-translator input.xlsx --target-lang ja

    # Translate directory with multiple target languages
    xlsx-translator --source-location ./excel_files --target-langs en,ja,es

    # Translate specific sheet and columns with business style
    xlsx-translator report.xlsx --sheet-name "Sales" --columns "A,C:E" --target-lang en --style-prompt business

    # Use context file for consistent terminology
    xlsx-translator technical.xlsx --target-lang ja --context-file glossary.json

    # Analyze file structure before translating
    xlsx-translator input.xlsx --analyze

    # Use smart context mode for better translations
    xlsx-translator data.xlsx --target-lang es --smart-context
"""

import argparse
import asyncio
import json
import logging
import os
import re
import shutil
import sys
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from copy import copy
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Optional, Set, Tuple, Union, Any
from collections import defaultdict

import google.generativeai as genai
import openpyxl
from colorama import Fore, Style, init
from openpyxl.utils import column_index_from_string, get_column_letter
from tqdm import tqdm

# Initialize colorama for colored output
init(autoreset=True)

# Constants
DEFAULT_BATCH_SIZE = 20
DEFAULT_CONCURRENCY = 4
MAX_RETRIES = 3
INVALID_SHEET_CHARS = r'[:\\/?*\[\]]'
MAX_SHEET_NAME_LENGTH = 31
CONTEXT_WINDOW_SIZE = 5  # Number of cells to include for context

# Style prompts
STYLE_PROMPTS = {
    'business': 'Translate using formal business language appropriate for professional documents.',
    'casual': 'Translate using conversational, friendly language suitable for informal communication.',
    'technical': 'Translate using precise technical terminology, maintaining accuracy for specialized content.',
    'marketing': 'Translate using persuasive, engaging language suitable for marketing materials.',
    'academic': 'Translate using scholarly language appropriate for academic or research contexts.',
    'legal': 'Translate using precise legal terminology, maintaining accuracy for legal documents.',
    'medical': 'Translate using accurate medical terminology for healthcare contexts.',
}


@dataclass
class TranslationConfig:
    """Configuration for translation job"""
    source_lang: str
    target_lang: str
    model_name: str
    batch_size: int
    concurrency: int
    preserve_formatting: bool
    preserve_formulas: bool
    style_prompt: str
    context_file: Optional[str]
    max_context_items: int
    debug: bool
    smart_context: bool = False
    analyze_only: bool = False


@dataclass
class CellFilter:
    """Filtering configuration for cells"""
    sheet_name: Optional[str]
    sheet_index: Optional[int]
    columns: Optional[Set[int]]
    rows: Optional[Set[int]]


@dataclass
class CellContext:
    """Context information for a cell"""
    row: int
    column: int
    value: str
    header: Optional[str] = None
    column_samples: List[str] = field(default_factory=list)
    row_context: List[str] = field(default_factory=list)
    cell_type: str = "text"  # text, number, formula, date, etc.


@dataclass
class SheetAnalysis:
    """Analysis results for a worksheet"""
    name: str
    total_cells: int
    text_cells: int
    formula_cells: int
    header_row: Optional[int]
    data_columns: List[int]
    sample_data: Dict[int, List[str]]


class ExcelAnalyzer:
    """Analyze Excel structure for better translation context"""
    
    @staticmethod
    def analyze_sheet(sheet) -> SheetAnalysis:
        """Analyze a worksheet to understand its structure"""
        analysis = SheetAnalysis(
            name=sheet.title,
            total_cells=0,
            text_cells=0,
            formula_cells=0,
            header_row=None,
            data_columns=[],
            sample_data={}
        )
        
        # Find header row (first row with multiple text cells)
        for row_idx in range(1, min(10, sheet.max_row + 1)):
            text_count = 0
            for col_idx in range(1, sheet.max_column + 1):
                cell = sheet.cell(row_idx, col_idx)
                if cell.value and isinstance(cell.value, str) and cell.value.strip():
                    text_count += 1
            
            if text_count >= 3:  # Likely a header row
                analysis.header_row = row_idx
                break
        
        # Analyze columns
        for col_idx in range(1, sheet.max_column + 1):
            col_data = []
            has_text = False
            
            for row_idx in range(1, min(100, sheet.max_row + 1)):
                cell = sheet.cell(row_idx, col_idx)
                if cell.value:
                    analysis.total_cells += 1
                    
                    if isinstance(cell.value, str) and cell.value.strip():
                        analysis.text_cells += 1
                        has_text = True
                        col_data.append(cell.value[:50])  # Sample
                    elif str(cell.value).startswith('='):
                        analysis.formula_cells += 1
            
            if has_text:
                analysis.data_columns.append(col_idx)
                analysis.sample_data[col_idx] = col_data[:5]  # Store samples
        
        return analysis
    
    @staticmethod
    def detect_table_structure(sheet) -> Dict[str, Any]:
        """Detect table structures in the sheet"""
        tables = []
        
        # Look for contiguous regions of data
        for row in range(1, min(50, sheet.max_row + 1)):
            for col in range(1, sheet.max_column + 1):
                cell = sheet.cell(row, col)
                if cell.value and isinstance(cell.value, str):
                    # Check if this could be a table header
                    table = ExcelAnalyzer._explore_table(sheet, row, col)
                    if table and table['cells'] > 9:  # Minimum table size
                        tables.append(table)
        
        return {'tables': tables}
    
    @staticmethod
    def _explore_table(sheet, start_row, start_col) -> Optional[Dict]:
        """Explore potential table starting from given position"""
        # Simple table detection - can be enhanced
        end_row = start_row
        end_col = start_col
        
        # Find table boundaries
        for r in range(start_row, min(start_row + 1000, sheet.max_row + 1)):
            if not sheet.cell(r, start_col).value:
                end_row = r - 1
                break
        else:
            end_row = min(start_row + 1000, sheet.max_row)
        
        for c in range(start_col, min(start_col + 50, sheet.max_column + 1)):
            if not sheet.cell(start_row, c).value:
                end_col = c - 1
                break
        else:
            end_col = min(start_col + 50, sheet.max_column)
        
        cells = (end_row - start_row + 1) * (end_col - start_col + 1)
        
        return {
            'start': (start_row, start_col),
            'end': (end_row, end_col),
            'cells': cells
        }


class ExcelTranslator:
    """Main translator class with enhanced context handling"""
    
    def __init__(self, config: TranslationConfig):
        self.config = config
        self.context_data = self._load_context()
        self._setup_logging()
        self._configure_api()
        self.analyzer = ExcelAnalyzer()
        
    def _setup_logging(self):
        """Configure logging based on debug flag"""
        level = logging.DEBUG if self.config.debug else logging.INFO
        logging.basicConfig(
            level=level,
            format='%(asctime)s - %(levelname)s - %(message)s'
        )
        self.logger = logging.getLogger(__name__)
        
    def _configure_api(self):
        """Configure Gemini API"""
        api_key = os.getenv('GOOGLE_API_KEY')
        if not api_key:
            raise ValueError(
                f"{Fore.RED}Error: GOOGLE_API_KEY environment variable not set.{Style.RESET_ALL}\n"
                f"{Fore.YELLOW}Please set your Gemini API key as an environment variable.{Style.RESET_ALL}"
            )
        genai.configure(api_key=api_key)
        self.model = genai.GenerativeModel(self.config.model_name)
        
    def _load_context(self) -> Dict:
        """Load context/glossary file if provided"""
        if not self.config.context_file:
            return {}
            
        try:
            with open(self.config.context_file, 'r', encoding='utf-8') as f:
                data = json.load(f)
                self.logger.info(f"Loaded context from {self.config.context_file}")
                return data
        except Exception as e:
            self.logger.warning(f"Failed to load context file: {e}")
            return {}
    
    def _build_prompt(self, cell_contexts: List[CellContext]) -> str:
        """Build translation prompt with enhanced context"""
        # Group cells by their context
        grouped_data = []
        
        for ctx in cell_contexts:
            cell_info = {
                "id": len(grouped_data),
                "text": ctx.value,
                "location": f"Row {ctx.row}, Column {get_column_letter(ctx.column)}"
            }
            
            # Add contextual information
            if ctx.header:
                cell_info["column_header"] = ctx.header
            
            if ctx.column_samples:
                cell_info["column_examples"] = ctx.column_samples[:3]
            
            if ctx.row_context:
                cell_info["row_context"] = ctx.row_context
            
            if ctx.cell_type != "text":
                cell_info["type"] = ctx.cell_type
            
            grouped_data.append(cell_info)
        
        context_str = ""
        if self.context_data:
            # Include relevant context items
            context_items = list(self.context_data.items())[:self.config.max_context_items]
            if context_items:
                context_str = "\nGlossary/Context:\n"
                for term, translation in context_items:
                    context_str += f"- {term} → {translation}\n"
        
        style_instruction = STYLE_PROMPTS.get(self.config.style_prompt, "")
        
        # Detect source language if set to auto
        source_lang_str = self.config.source_lang
        if source_lang_str == "auto":
            source_lang_str = "the source language (auto-detect)"
        
        prompt = f"""
You are translating an Excel spreadsheet from {source_lang_str} to {self.config.target_lang}.
{style_instruction}
{context_str}

IMPORTANT INSTRUCTIONS:
1. Maintain ALL formatting EXACTLY (spaces, punctuation, capitalization patterns)
2. For empty strings, return empty strings
3. For numbers only, return the same numbers
4. Use the column headers and examples to understand context
5. Apply glossary terms consistently
6. Consider the cell's position and surrounding context
7. If a cell contains a formula reference (like "Total:", "Sum:", etc.), translate appropriately

The input contains cells with their context. Use this context to produce accurate, contextually appropriate translations.

Return ONLY valid JSON in this exact format:
{{
    "translations": [
        {{"id": <id>, "translation": "<translated_text>"}}
    ]
}}

Input data:
{json.dumps(grouped_data, ensure_ascii=False, indent=2)}
"""
        return prompt
    
    def _extract_json(self, response_text: str) -> str:
        """Extract JSON from response"""
        # Try to find JSON code blocks
        matches = re.findall(r'```(?:json)?\n(.*?)\n```', response_text, re.DOTALL)
        if matches:
            return matches[0]
        
        # Try to find raw JSON
        match = re.search(r'{.*}', response_text, re.DOTALL)
        if match:
            return match.group(0)
            
        return response_text
    
    def _get_cell_context(self, sheet, row: int, col: int, 
                         headers: Dict[int, str], 
                         column_data: Dict[int, List]) -> CellContext:
        """Extract context for a cell"""
        cell = sheet.cell(row, col)
        context = CellContext(
            row=row,
            column=col,
            value=str(cell.value) if cell.value else ""
        )
        
        # Add header if available
        context.header = headers.get(col)
        
        # Add column samples (excluding current cell)
        if col in column_data:
            samples = [str(v) for v in column_data[col] if v != cell.value]
            context.column_samples = samples[:3]
        
        # Add row context (neighboring cells)
        row_context = []
        for offset in [-2, -1, 1, 2]:
            neighbor_col = col + offset
            if 1 <= neighbor_col <= sheet.max_column:
                neighbor = sheet.cell(row, neighbor_col)
                if neighbor.value:
                    row_context.append(str(neighbor.value)[:30])
        context.row_context = row_context
        
        # Determine cell type
        if isinstance(cell.value, (int, float)):
            context.cell_type = "number"
        elif cell.value and str(cell.value).startswith('='):
            context.cell_type = "formula"
        elif cell.is_date:
            context.cell_type = "date"
        
        return context
    
    def translate_batch_with_context(self, contexts: List[CellContext]) -> Dict[int, str]:
        """Translate a batch of cells with their context"""
        if not contexts:
            return {}
            
        for attempt in range(MAX_RETRIES):
            try:
                prompt = self._build_prompt(contexts)
                response = self.model.generate_content(prompt)
                
                if self.config.debug:
                    self.logger.debug(f"API Response: {response.text[:200]}...")
                
                cleaned = self._extract_json(response.text)
                result = json.loads(cleaned)
                
                if 'translations' not in result:
                    raise ValueError("Missing 'translations' key in response")
                
                # Build translation map
                translation_map = {}
                for item in result['translations']:
                    idx = item.get('id')
                    translation = item.get('translation', '')
                    if idx is not None:
                        translation_map[idx] = translation
                
                return translation_map
                
            except Exception as e:
                self.logger.warning(f"Translation attempt {attempt + 1} failed: {e}")
                if attempt < MAX_RETRIES - 1:
                    time.sleep(2 ** attempt)  # Exponential backoff
                else:
                    self.logger.error(f"All translation attempts failed for batch")
                    return {i: ctx.value for i, ctx in enumerate(contexts)}
    
    def analyze_file(self, input_path: Path) -> Dict[str, Any]:
        """Analyze Excel file structure"""
        print(f"\n{Fore.CYAN}Analyzing {input_path}...{Style.RESET_ALL}")
        
        wb = openpyxl.load_workbook(input_path, data_only=True)
        analysis_results = {
            'file': str(input_path),
            'sheets': []
        }
        
        for sheet in wb.worksheets:
            sheet_analysis = self.analyzer.analyze_sheet(sheet)
            table_info = self.analyzer.detect_table_structure(sheet)
            
            analysis_results['sheets'].append({
                'name': sheet_analysis.name,
                'total_cells': sheet_analysis.total_cells,
                'text_cells': sheet_analysis.text_cells,
                'formula_cells': sheet_analysis.formula_cells,
                'header_row': sheet_analysis.header_row,
                'text_columns': len(sheet_analysis.data_columns),
                'tables_found': len(table_info['tables']),
                'sample_headers': [sheet.cell(sheet_analysis.header_row or 1, col).value 
                                 for col in sheet_analysis.data_columns[:5]]
                                 if sheet_analysis.header_row else []
            })
        
        wb.close()
        
        # Print analysis
        print(f"\n{Fore.GREEN}File Analysis Results:{Style.RESET_ALL}")
        print(f"File: {analysis_results['file']}")
        print(f"Sheets: {len(analysis_results['sheets'])}")
        
        for sheet_info in analysis_results['sheets']:
            print(f"\n  Sheet: {Fore.YELLOW}{sheet_info['name']}{Style.RESET_ALL}")
            print(f"    Total cells: {sheet_info['total_cells']}")
            print(f"    Text cells: {sheet_info['text_cells']}")
            print(f"    Formula cells: {sheet_info['formula_cells']}")
            print(f"    Text columns: {sheet_info['text_columns']}")
            if sheet_info['sample_headers']:
                print(f"    Headers: {', '.join(str(h) for h in sheet_info['sample_headers'][:5])}")
        
        return analysis_results
    
    def translate_file(self, input_path: Path, output_path: Path, cell_filter: CellFilter):
        """Translate a single Excel file"""
        self.logger.info(f"Processing: {input_path}")
        
        # Analyze file first if requested
        if self.config.analyze_only:
            self.analyze_file(input_path)
            return
        
        try:
            # Load workbook
            wb = openpyxl.load_workbook(input_path, data_only=False, keep_vba=True)
            
            # Create temporary translated workbook
            temp_file = output_path.with_suffix('.tmp' + output_path.suffix)
            wb_trans = openpyxl.Workbook()
            if 'Sheet' in wb_trans.sheetnames:
                del wb_trans['Sheet']
            
            # Get sheets to process
            sheets_to_process = self._get_sheets_to_process(wb, cell_filter)
            
            # Process each sheet
            for sheet_name in tqdm(sheets_to_process, desc="Sheets", leave=False):
                if self.config.smart_context:
                    self._process_sheet_smart(wb[sheet_name], wb_trans, cell_filter)
                else:
                    self._process_sheet(wb[sheet_name], wb_trans, cell_filter)
            
            # Save translated workbook
            wb_trans.save(temp_file)
            wb_trans.close()
            wb.close()
            
            # Apply formatting preservation
            if self.config.preserve_formatting:
                self._preserve_formatting(input_path, temp_file, output_path)
                temp_file.unlink()  # Remove temp file
            else:
                shutil.move(temp_file, output_path)
            
            self.logger.info(f"✅ Created: {output_path}")
            
        except Exception as e:
            self.logger.error(f"Failed to process {input_path}: {e}")
            raise
    
    def _get_sheets_to_process(self, wb: openpyxl.Workbook, cell_filter: CellFilter) -> List[str]:
        """Determine which sheets to process based on filter"""
        if cell_filter.sheet_name:
            if cell_filter.sheet_name in wb.sheetnames:
                return [cell_filter.sheet_name]
            else:
                raise ValueError(f"Sheet '{cell_filter.sheet_name}' not found")
        elif cell_filter.sheet_index is not None:
            if 0 <= cell_filter.sheet_index < len(wb.sheetnames):
                return [wb.sheetnames[cell_filter.sheet_index]]
            else:
                raise ValueError(f"Sheet index {cell_filter.sheet_index} out of range")
        else:
            return wb.sheetnames
    
    def _process_sheet(self, source_sheet, target_wb, cell_filter: CellFilter):
        """Process a single worksheet (standard mode)"""
        # Translate sheet name
        sheet_name_trans = self.translate_batch_with_context(
            [CellContext(0, 0, source_sheet.title)]
        )
        new_name = sheet_name_trans.get(0, source_sheet.title)
        new_name = re.sub(INVALID_SHEET_CHARS, '_', new_name)[:MAX_SHEET_NAME_LENGTH]
        
        target_sheet = target_wb.create_sheet(new_name)
        
        # Copy sheet properties
        target_sheet.sheet_properties = copy(source_sheet.sheet_properties)
        
        # Collect all texts to translate
        texts_to_translate = []
        cell_mapping = {}
        
        for row in source_sheet.iter_rows():
            # Check row filter
            if cell_filter.rows and row[0].row not in cell_filter.rows:
                continue
                
            for cell in row:
                # Check column filter
                if cell_filter.columns and cell.column not in cell_filter.columns:
                    continue
                
                # Copy cell to target
                target_cell = target_sheet.cell(row=cell.row, column=cell.column)
                
                # Copy style if preserving formatting
                if self.config.preserve_formatting and cell.has_style:
                    target_cell.font = copy(cell.font)
                    target_cell.border = copy(cell.border)
                    target_cell.fill = copy(cell.fill)
                    target_cell.number_format = cell.number_format
                    target_cell.alignment = copy(cell.alignment)
                
                # Handle cell value
                if cell.value is None:
                    continue
                elif self.config.preserve_formulas and str(cell.value).startswith('='):
                    target_cell.value = cell.value  # Keep formula
                elif isinstance(cell.value, str) and cell.value.strip():
                    # Add to translation batch
                    batch_id = len(texts_to_translate)
                    texts_to_translate.append(
                        CellContext(cell.row, cell.column, cell.value)
                    )
                    cell_mapping[batch_id] = (cell.row, cell.column)
                else:
                    target_cell.value = cell.value  # Copy as-is
        
        # Translate in batches
        self._translate_cells_concurrent(texts_to_translate, cell_mapping, target_sheet)
        
        # Copy other sheet properties
        self._copy_sheet_properties(source_sheet, target_sheet)
    
    def _process_sheet_smart(self, source_sheet, target_wb, cell_filter: CellFilter):
        """Process worksheet with smart context extraction"""
        # First, analyze the sheet
        analysis = self.analyzer.analyze_sheet(source_sheet)
        
        # Extract headers
        headers = {}
        if analysis.header_row:
            for col in range(1, source_sheet.max_column + 1):
                header_cell = source_sheet.cell(analysis.header_row, col)
                if header_cell.value:
                    headers[col] = str(header_cell.value)
        
        # Extract column data for context
        column_data = defaultdict(list)
        for col in analysis.data_columns:
            for row in range(1, min(20, source_sheet.max_row + 1)):
                cell = source_sheet.cell(row, col)
                if cell.value and isinstance(cell.value, str):
                    column_data[col].append(cell.value)
        
        # Translate sheet name
        sheet_name_trans = self.translate_batch_with_context(
            [CellContext(0, 0, source_sheet.title)]
        )
        new_name = sheet_name_trans.get(0, source_sheet.title)
        new_name = re.sub(INVALID_SHEET_CHARS, '_', new_name)[:MAX_SHEET_NAME_LENGTH]
        
        target_sheet = target_wb.create_sheet(new_name)
        target_sheet.sheet_properties = copy(source_sheet.sheet_properties)
        
        # Collect cells with enhanced context
        contexts_to_translate = []
        cell_mapping = {}
        
        for row in source_sheet.iter_rows():
            if cell_filter.rows and row[0].row not in cell_filter.rows:
                continue
                
            for cell in row:
                if cell_filter.columns and cell.column not in cell_filter.columns:
                    continue
                
                # Copy cell to target
                target_cell = target_sheet.cell(row=cell.row, column=cell.column)
                
                # Copy style
                if self.config.preserve_formatting and cell.has_style:
                    target_cell.font = copy(cell.font)
                    target_cell.border = copy(cell.border)
                    target_cell.fill = copy(cell.fill)
                    target_cell.number_format = cell.number_format
                    target_cell.alignment = copy(cell.alignment)
                
                # Handle cell value
                if cell.value is None:
                    continue
                elif self.config.preserve_formulas and str(cell.value).startswith('='):
                    target_cell.value = cell.value
                elif isinstance(cell.value, str) and cell.value.strip():
                    # Create context-rich cell info
                    context = self._get_cell_context(
                        source_sheet, cell.row, cell.column, headers, column_data
                    )
                    batch_id = len(contexts_to_translate)
                    contexts_to_translate.append(context)
                    cell_mapping[batch_id] = (cell.row, cell.column)
                else:
                    target_cell.value = cell.value
        
        # Translate with context
        self._translate_cells_concurrent(contexts_to_translate, cell_mapping, target_sheet)
        
        # Copy sheet properties
        self._copy_sheet_properties(source_sheet, target_sheet)
    
    def _translate_cells_concurrent(self, contexts: List[CellContext], 
                                   cell_mapping: Dict[int, Tuple[int, int]], 
                                   target_sheet):
        """Translate cells using concurrent API calls"""
        if not contexts:
            return
            
        # Group related cells for better context
        batches = []
        if self.config.smart_context:
            # Group by column for better context
            column_groups = defaultdict(list)
            for i, ctx in enumerate(contexts):
                column_groups[ctx.column].append((i, ctx))
            
            # Create batches maintaining column grouping
            current_batch = []
            for col, items in sorted(column_groups.items()):
                for i, ctx in items:
                    current_batch.append((i, ctx))
                    if len(current_batch) >= self.config.batch_size:
                        batches.append(current_batch)
                        current_batch = []
            
            if current_batch:
                batches.append(current_batch)
        else:
            # Simple batching
            for i in range(0, len(contexts), self.config.batch_size):
                batch = [(j, contexts[j]) for j in range(i, min(i + self.config.batch_size, len(contexts)))]
                batches.append(batch)
        
        # Process batches concurrently
        all_translations = {}
        with ThreadPoolExecutor(max_workers=self.config.concurrency) as executor:
            future_to_batch = {
                executor.submit(
                    self.translate_batch_with_context, 
                    [ctx for _, ctx in batch]
                ): batch 
                for batch in batches
            }
            
            for future in tqdm(as_completed(future_to_batch), 
                             total=len(batches), 
                             desc="Translating", 
                             leave=False):
                batch = future_to_batch[future]
                translations = future.result()
                
                # Map back to original indices
                for idx, (orig_idx, _) in enumerate(batch):
                    if idx in translations:
                        all_translations[orig_idx] = translations[idx]
        
        # Apply translations to cells
        for batch_id, translation in all_translations.items():
            if batch_id in cell_mapping:
                row, col = cell_mapping[batch_id]
                target_sheet.cell(row=row, column=col).value = translation
    
    def _copy_sheet_properties(self, source_sheet, target_sheet):
        """Copy sheet properties like column widths, row heights, etc."""
        # Copy column dimensions
        for col, dim in source_sheet.column_dimensions.items():
            target_sheet.column_dimensions[col] = copy(dim)
        
        # Copy row dimensions
        for row, dim in source_sheet.row_dimensions.items():
            target_sheet.row_dimensions[row] = copy(dim)
        
        # Copy print settings
        target_sheet.print_options = copy(source_sheet.print_options)
        target_sheet.page_setup = copy(source_sheet.page_setup)
        target_sheet.page_margins = copy(source_sheet.page_margins)
        
        # Copy freeze panes
        target_sheet.freeze_panes = source_sheet.freeze_panes
    
    def _preserve_formatting(self, original_file: Path, translated_file: Path, output_file: Path):
        """Preserve formatting by copying translated values to original structure"""
        try:
            # Copy original to output
            shutil.copy2(original_file, output_file)
            time.sleep(0.1)  # Brief pause for file system
            
            # Open files
            wb_trans = openpyxl.load_workbook(translated_file, data_only=True)
            wb_final = openpyxl.load_workbook(output_file, keep_vba=True)
            
            # Copy translated content
            for idx, sheet_name in enumerate(wb_trans.sheetnames):
                if idx < len(wb_final.worksheets):
                    trans_sheet = wb_trans.worksheets[idx]
                    final_sheet = wb_final.worksheets[idx]
                    
                    # Update sheet name
                    try:
                        final_sheet.title = trans_sheet.title
                    except:
                        pass  # Keep original if rename fails
                    
                    # Handle merged cells
                    merged_ranges = list(final_sheet.merged_cells.ranges)
                    for merged_range in merged_ranges:
                        final_sheet.unmerge_cells(str(merged_range))
                    
                    # Copy values
                    for row in trans_sheet.iter_rows():
                        for cell in row:
                            if cell.value is not None:
                                final_sheet.cell(row=cell.row, column=cell.column).value = cell.value
                    
                    # Re-merge cells
                    for merged_range in merged_ranges:
                        final_sheet.merge_cells(str(merged_range))
            
            # Save final file
            wb_final.save(output_file)
            wb_final.close()
            wb_trans.close()
            
        except Exception as e:
            self.logger.error(f"Format preservation failed: {e}")
            # Fall back to translated file
            shutil.copy2(translated_file, output_file)


def parse_columns(columns_str: str) -> Set[int]:
    """Parse column specification like 'A,C:E,G' into column indices"""
    columns = set()
    for part in columns_str.split(','):
        part = part.strip()
        if ':' in part:
            # Range like C:E
            start, end = part.split(':')
            start_idx = column_index_from_string(start)
            end_idx = column_index_from_string(end)
            columns.update(range(start_idx, end_idx + 1))
        else:
            # Single column
            columns.add(column_index_from_string(part))
    return columns


def parse_rows(rows_str: str) -> Set[int]:
    """Parse row specification like '1,5-10,15' into row indices"""
    rows = set()
    for part in rows_str.split(','):
        part = part.strip()
        if '-' in part:
            # Range like 5-10
            start, end = map(int, part.split('-'))
            rows.update(range(start, end + 1))
        else:
            # Single row
            rows.add(int(part))
    return rows


def process_directory(translator: ExcelTranslator, source_dir: Path, output_dir: Path, 
                     cell_filter: CellFilter, target_langs: List[str]):
    """Process all Excel files in a directory"""
    excel_files = list(source_dir.glob('*.xlsx')) + list(source_dir.glob('*.xlsm'))
    
    if not excel_files:
        print(f"{Fore.YELLOW}No Excel files found in {source_dir}{Style.RESET_ALL}")
        return
    
    print(f"{Fore.GREEN}Found {len(excel_files)} Excel files{Style.RESET_ALL}")
    
    for target_lang in target_langs:
        print(f"\n{Fore.CYAN}Translating to {target_lang}{Style.RESET_ALL}")
        
        # Update translator config
        translator.config.target_lang = target_lang
        
        # Create language subdirectory
        lang_dir = output_dir / target_lang
        lang_dir.mkdir(parents=True, exist_ok=True)
        
        # Process each file
        for excel_file in tqdm(excel_files, desc=f"Files ({target_lang})"):
            output_name = f"{excel_file.stem}_{target_lang}{excel_file.suffix}"
            output_path = lang_dir / output_name
            
            try:
                translator.translate_file(excel_file, output_path, cell_filter)
            except Exception as e:
                print(f"{Fore.RED}Failed to process {excel_file}: {e}{Style.RESET_ALL}")


def print_supported_languages():
    """Print information about language support"""
    print(f"\n{Fore.CYAN}Language Support:{Style.RESET_ALL}")
    print(f"{Fore.GREEN}This tool supports any language that the Gemini model can handle.{Style.RESET_ALL}")
    print(f"\nCommon language codes include:")
    
    common_langs = [
        ('en', 'English'), ('es', 'Spanish'), ('fr', 'French'),
        ('de', 'German'), ('it', 'Italian'), ('pt', 'Portuguese'),
        ('ru', 'Russian'), ('ja', 'Japanese'), ('ko', 'Korean'),
        ('zh', 'Chinese (Simplified)'), ('zh-TW', 'Chinese (Traditional)'),
        ('ar', 'Arabic'), ('hi', 'Hindi'), ('tr', 'Turkish'),
        ('pl', 'Polish'), ('nl', 'Dutch'), ('sv', 'Swedish'),
        ('da', 'Danish'), ('no', 'Norwegian'), ('fi', 'Finnish'),
        ('he', 'Hebrew'), ('th', 'Thai'), ('vi', 'Vietnamese'),
        ('id', 'Indonesian'), ('ms', 'Malay'), ('cs', 'Czech'),
        ('hu', 'Hungarian'), ('ro', 'Romanian'), ('bg', 'Bulgarian'),
        ('uk', 'Ukrainian'), ('el', 'Greek'), ('fa', 'Persian'),
    ]
    
    for i in range(0, len(common_langs), 3):
        row_langs = common_langs[i:i+3]
        row_str = "  "
        for code, name in row_langs:
            row_str += f"{Fore.GREEN}{code:<8}{Style.RESET_ALL}{name:<25}"
        print(row_str)
    
    print(f"\n{Fore.YELLOW}Note: You can use any ISO 639-1 or ISO 639-2 language code.{Style.RESET_ALL}")
    print(f"The model will attempt to translate between any language pair it knows.")


def main():
    """Main entry point"""
    parser = argparse.ArgumentParser(
        description="Excel Translation Utility with Gemini API",
        epilog="Examples:\n"
               "  %(prog)s input.xlsx --target-lang ja\n"
               "  %(prog)s --source-location ./files --target-langs en,ja,es\n"
               "  %(prog)s report.xlsx --sheet-name Sales --columns A,C:E --target-lang en\n"
               "  %(prog)s data.xlsx --analyze\n"
               "  %(prog)s data.xlsx --target-lang es --smart-context",
        formatter_class=argparse.RawDescriptionHelpFormatter
    )
    
    # Input options (mutually exclusive)
    input_group = parser.add_mutually_exclusive_group(required=False)
    input_group.add_argument('input_file', nargs='?', help='Input Excel file (.xlsx or .xlsm)')
    input_group.add_argument('--source-location', type=str, help='Directory containing Excel files')
    
    # Language options
    parser.add_argument('--source-lang', default='auto', 
                       help='Source language code (default: auto-detect)')
    
    lang_group = parser.add_mutually_exclusive_group()
    lang_group.add_argument('--target-lang', help='Target language code (e.g., ja)')
    lang_group.add_argument('--target-langs', help='Multiple target languages (comma-separated)')
    
    # Excel-specific options
    parser.add_argument('--sheet-name', help='Process only this sheet by name')
    parser.add_argument('--sheet-index', type=int, help='Process only this sheet by index (0-based)')
    parser.add_argument('--columns', help='Columns to translate (e.g., "A,C:E,G")')
    parser.add_argument('--rows', help='Rows to translate (e.g., "1,5-10,15")')
    parser.add_argument('--preserve-formatting', action='store_true', default=True,
                       help='Preserve cell formatting (default: True)')
    parser.add_argument('--preserve-formulas', action='store_true', default=True,
                       help='Skip cells containing formulas (default: True)')
    
    # Context and styling
    parser.add_argument('--context-file', help='Path to glossary/context JSON file')
    parser.add_argument('--style-prompt', choices=list(STYLE_PROMPTS.keys()), 
                       default='business', help='Translation style')
    parser.add_argument('--max-context-items', type=int, default=10,
                       help='Maximum context items to include')
    parser.add_argument('--smart-context', action='store_true',
                       help='Use enhanced context extraction for better translations')
    
    # Output and processing
    parser.add_argument('--output-dir', default='.', help='Output directory')
    parser.add_argument('--model', default='gemini-2.0-flash', help='Gemini model name')
    parser.add_argument('--batch-size', type=int, default=DEFAULT_BATCH_SIZE,
                       help=f'Cells per API call (default: {DEFAULT_BATCH_SIZE})')
    parser.add_argument('--concurrency', type=int, default=DEFAULT_CONCURRENCY,
                       help=f'Concurrent API calls (default: {DEFAULT_CONCURRENCY})')
    
    # Info flags
    parser.add_argument('--list-styles', action='store_true', help='List available style options')
    parser.add_argument('--list-languages', action='store_true', help='Show language code information')
    parser.add_argument('--analyze', action='store_true', help='Analyze file structure without translating')
    parser.add_argument('--debug', action='store_true', help='Enable debug logging')
    
    args = parser.parse_args()
    
    # Handle info flags
    if args.list_styles:
        print(f"\n{Fore.CYAN}Available style prompts:{Style.RESET_ALL}")
        for style, description in STYLE_PROMPTS.items():
            print(f"  {Fore.GREEN}{style:<12}{Style.RESET_ALL} - {description}")
        return
    
    if args.list_languages:
        print_supported_languages()
        return
    
    # Validate input
    if not args.input_file and not args.source_location:
        parser.error("Either input_file or --source-location is required")
    
    if not args.target_lang and not args.target_langs and not args.analyze:
        parser.error("Either --target-lang, --target-langs, or --analyze is required")
    
    # Parse target languages
    target_langs = []
    if args.target_langs:
        target_langs = [lang.strip() for lang in args.target_langs.split(',')]
    elif args.target_lang:
        target_langs = [args.target_lang]
    
    # Create configuration
    config = TranslationConfig(
        source_lang=args.source_lang,
        target_lang=target_langs[0] if target_langs else '',
        model_name=args.model,
        batch_size=args.batch_size,
        concurrency=args.concurrency,
        preserve_formatting=args.preserve_formatting,
        preserve_formulas=args.preserve_formulas,
        style_prompt=args.style_prompt,
        context_file=args.context_file,
        max_context_items=args.max_context_items,
        debug=args.debug,
        smart_context=args.smart_context,
        analyze_only=args.analyze
    )
    
    # Create cell filter
    cell_filter = CellFilter(
        sheet_name=args.sheet_name,
        sheet_index=args.sheet_index,
        columns=parse_columns(args.columns) if args.columns else None,
        rows=parse_rows(args.rows) if args.rows else None
    )
    
    # Create translator
    try:
        translator = ExcelTranslator(config)
    except ValueError as e:
        print(str(e))
        sys.exit(1)
    
    # Process files
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    
    try:
        if args.input_file:
            # Single file mode
            input_path = Path(args.input_file)
            if not input_path.exists():
                print(f"{Fore.RED}Error: File not found: {input_path}{Style.RESET_ALL}")
                sys.exit(1)
            
            if args.analyze:
                translator.analyze_file(input_path)
            else:
                for target_lang in target_langs:
                    translator.config.target_lang = target_lang
                    output_name = f"{input_path.stem}_{target_lang}{input_path.suffix}"
                    output_path = output_dir / output_name
                    
                    print(f"{Fore.CYAN}Translating to {target_lang}...{Style.RESET_ALL}")
                    translator.translate_file(input_path, output_path, cell_filter)
        
        else:
            # Directory mode
            source_dir = Path(args.source_location)
            if not source_dir.exists():
                print(f"{Fore.RED}Error: Directory not found: {source_dir}{Style.RESET_ALL}")
                sys.exit(1)
            
            if args.analyze:
                # Analyze all files in directory
                excel_files = list(source_dir.glob('*.xlsx')) + list(source_dir.glob('*.xlsm'))
                for excel_file in excel_files:
                    translator.analyze_file(excel_file)
            else:
                process_directory(translator, source_dir, output_dir, cell_filter, target_langs)
        
        if not args.analyze:
            print(f"\n{Fore.GREEN}✅ Translation complete!{Style.RESET_ALL}")
        
    except Exception as e:
        print(f"\n{Fore.RED}❌ Translation failed: {e}{Style.RESET_ALL}")
        if args.debug:
            import traceback
            traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()