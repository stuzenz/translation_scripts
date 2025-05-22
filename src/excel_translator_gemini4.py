"""
Excel Translation Toolkit with Format Preservation - Gemini API Version

Features:
1. Full Excel file translation using Gemini API models
2. Automatic format preservation (styles, images, merged cells, drawings)
3. Macro-enabled file (.xlsm) support
4. Batch processing with progress tracking
5. Controlled JSON response format from Gemini API

Usage Modes:
1. Full translation with post-processing (default)

Command Line Interface:
python excel_translator_gemini.py translate <input_file> [options]
"""

import argparse
import openpyxl
from openpyxl.utils import get_column_letter
import os
import re
import logging
from copy import copy, deepcopy
from tqdm import tqdm
import shutil
import subprocess
import json
import google.generativeai as genai
from colorama import Fore, Style, init
import time # Import time for potential wait

# Initialize colorama
init(autoreset=True)

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# Constants
MAX_RETRIES = 2
DEFAULT_BATCH_SIZE = 15 # Adjusted batch size to be closer to docx script recommendation
INVALID_SHEET_CHARS = r'[:\\/?*\[\]]'

def install_dependencies():
    """Install required Python packages."""
    try:
        import openpyxl
        from tqdm import tqdm
        import google.generativeai
        from colorama import Fore, Style, init
    except ImportError:
        print("Installing dependencies...")
        subprocess.check_call(["pip", "install", "openpyxl", "tqdm", "google-generativeai", "colorama"])
        print("Dependencies installed.")

def extract_json(response_text):
    """Extract JSON from markdown code blocks or raw text (same as docx script)"""
    # Try to find JSON code blocks
    matches = re.findall(r'```(?:json)?\n(.*?)\n```', response_text, re.DOTALL)
    if matches:
        return matches[0]
    # If no code blocks, try to find first JSON structure
    match = re.search(r'{(.*?)}', response_text, re.DOTALL)
    if match:
        return f'{{{match.group(1)}}}'
    return response_text

# Translation Core Functions
def translate_text_with_gemini(texts, source_lang="auto", target_lang="en", model_name="gemini-1.5-flash"):
    """Robust batch translation with Gemini API and JSON response handling"""
    if not texts:
        return []

    genai.configure(api_key=os.getenv('GOOGLE_API_KEY')) # Ensure API key is configured here as well
    model = genai.GenerativeModel(model_name)

    batch_data = [{"id": i, "text": text} for i, text in enumerate(texts)]
    prompt = f"""
        Translate from {source_lang} to {target_lang}. Maintain formatting EXACTLY.
        Return ONLY VALID JSON using this format:
        {{
            "translations": [
                {{"id": <original_id>, "translation": "<translated_text>"}}
            ]
        }}
        DO NOT USE MARKDOWN. Ensure proper JSON escaping.
        Input: {json.dumps(batch_data, ensure_ascii=False)}
        """

    try:
        response = model.generate_content(prompt)
        print(f"{Fore.GREEN}📥 Raw Response:{Style.RESET_ALL} {response.text[:150]}...")

        cleaned_response = extract_json(response.text)
        print(f"{Fore.BLUE}🧹 Cleaned Response:{Style.RESET_ALL} {cleaned_response[:150]}...")

        try:
            result = json.loads(cleaned_response)
        except json.JSONDecodeError as e:
            print(f"{Fore.RED}❌ JSON Error:{Style.RESET_ALL} {str(e)}")
            print(f"{Fore.MAGENTA}🧬 Response Fragment:{Style.RESET_ALL} {cleaned_response[:500]}")
            return [None] * len(texts)  # Return None for all in batch on JSON error

        if 'translations' not in result:
            print(f"{Fore.RED}⚠️ Missing 'translations' key in JSON response{Style.RESET_ALL}")
            return [None] * len(texts) # Return None for all in batch if translations key missing

        translated_texts = [None] * len(texts) # Initialize list with None values
        for item in result['translations']:
            item_id = item.get('id')
            translation = item.get('translation')
            if item_id is not None and 0 <= item_id < len(texts): # Basic validation
                translated_texts[item_id] = translation

        return translated_texts

    except Exception as e:
        logging.error(f"Translation error: {e}")
        print(f"{Fore.RED}💥 Gemini API Error:{Style.RESET_ALL} {str(e)}")
        return [None] * len(texts) # Return None for all in batch on API error


# Format Preservation Functions (Modified to use provided logic)
def copy_translated_values(source_sheet, target_sheet):
    """Copy values from translated sheet to target sheet while preserving formatting"""
    for row in source_sheet.iter_rows():
        for cell in row:
            # Only copy value if cell is not empty in source
            if cell.value is not None:
                target_cell = target_sheet.cell(row=cell.row, column=cell.column)
                target_cell.value = cell.value

def process_merged_cells(sheet, translated_sheet):
    """Handle merged cells by unmerging, copying values, then re-merging"""
    # Record original merged ranges
    merged_ranges = [str(r) for r in sheet.merged_cells.ranges]

    # Unmerge all cells
    for merged_range in merged_ranges:
        sheet.unmerge_cells(merged_range)

    # Copy values from translated sheet
    copy_translated_values(translated_sheet, sheet)

    # Re-apply original merged ranges
    for merged_range in merged_ranges:
        sheet.merge_cells(merged_range)


def preserve_formatting(original_file, translated_file, output_file_ext): # Added output_file_ext
    """Main format preservation routine - using provided script logic"""
    base_name = os.path.splitext(original_file)[0]
    final_file = f"{base_name}_ja{output_file_ext}" # Use dynamic extension

    try:
        # Create working copy of original file
        if not os.path.exists(final_file):
            shutil.copyfile(original_file, final_file)
            time.sleep(0.5) # Small wait after file copy

        # Open workbooks
        translated_wb = openpyxl.load_workbook(translated_file)
        final_wb = openpyxl.load_workbook(final_file, keep_vba=True) # Keep VBA for macro files if needed

        # Process each sheet by index
        for sheet_idx in tqdm(range(len(translated_wb.sheetnames)), desc="Processing sheets"):
            if sheet_idx >= len(final_wb.sheetnames):
                break  # Handle sheet count mismatch

            translated_sheet = translated_wb.worksheets[sheet_idx]
            final_sheet = final_wb.worksheets[sheet_idx]

            # Update sheet name if different
            if translated_sheet.title != final_sheet.title:
                try:
                    final_sheet.title = translated_sheet.title[:31]  # Excel max 31 chars
                except Exception as e:
                    logging.warning(f"Couldn't rename sheet: {e}")

            # Process merged cells and copy values
            process_merged_cells(final_sheet, translated_sheet)

        # Preserve other workbook properties - Dimensions (already handled in process_sheet - no need to repeat here)

        final_wb.save(final_file)
        logging.info(f"Successfully created final file: {final_file}")
        return True

    except Exception as e:
        logging.error(f"Post-processing failed: {e}")
        return False


def process_sheet(orig_sheet, trans_sheet, final_sheet):
    """Process individual sheet - now includes drawing copy"""
    try:
        # Copy dimensions
        for col, dim in orig_sheet.column_dimensions.items():
            final_sheet.column_dimensions[col] = copy(dim)
        for row, dim in orig_sheet.row_dimensions.items():
            final_sheet.row_dimensions[row] = copy(dim)

        # Copy drawings - IMPORTANT: Deepcopy drawings to avoid corruption and broken links
        if orig_sheet.drawing:
            final_sheet.drawing = deepcopy(orig_sheet.drawing)

    except Exception as e:
        logging.error(f"Sheet processing error: {e}")


def copy_all_values(source_sheet, target_sheet): # Redundant now - using copy_translated_values in post_process
    """Deep copy of all cell values - Redundant, using copy_translated_values instead"""
    for row in source_sheet.iter_rows():
        for cell in row:
            target_sheet.cell(row=cell.row, column=cell.column).value = cell.value

# Main Translation Processor
def translate_excel(input_file, source_lang, target_lang, model_name, batch_size, post_process):
    """End-to-end translation workflow using Gemini API"""
    try:
        wb = openpyxl.load_workbook(input_file)
        base_name = os.path.splitext(input_file)[0]
        input_file_ext = os.path.splitext(input_file)[1] # Get extension
        output_file = f"{base_name}_translated{input_file_ext}" # Use input file extension

        # Create translated workbook
        wb_out = openpyxl.Workbook()
        del wb_out['Sheet']

        # Translate sheet names
        sheet_names = {}
        for sheet in tqdm(wb.sheetnames, desc="Translating Sheets"):
            translated = translate_text_with_gemini([sheet], source_lang, target_lang, model_name)
            clean_name = re.sub(INVALID_SHEET_CHARS, '_', translated[0] or sheet)[:31] # Get first element of list
            sheet_names[sheet] = clean_name

        # Process content
        for sheet_name in tqdm(wb.sheetnames, desc="Processing Content"):
            orig_sheet = wb[sheet_name]
            trans_sheet = wb_out.create_sheet(sheet_names[sheet_name])
            process_worksheet(orig_sheet, trans_sheet, source_lang, target_lang, model_name, batch_size)
            process_sheet(orig_sheet, trans_sheet, trans_sheet) # Call process_sheet to copy drawings and dimensions

        wb_out.save(output_file)
        logging.info(f"Translation complete: {output_file}")

        post_process_success = True # Assume success unless post_process fails
        if post_process:
            post_process_success = preserve_formatting(input_file, output_file, input_file_ext) # Pass extension

        if post_process_success: # Only delete if post-process or no post-process was successful
            try:
                os.remove(output_file) # Remove intermediary file
                logging.info(f"Removed intermediary file: {output_file}")
            except OSError as e:
                logging.warning(f"Error deleting intermediary file {output_file}: {e}")

        return post_process_success # Return if post processing was successful, or translation if no post processing

    except Exception as e:
        logging.error(f"Translation failed: {e}")
        return False

def process_worksheet(source_sheet, target_sheet, src_lang, tgt_lang, model_name, batch_size):
    """Process individual worksheet using Gemini API - content and styles, drawing copy moved to process_sheet"""
    # Copy styles and content, but not drawings - drawings are handled in process_sheet

    # Copy styles
    for row in source_sheet.iter_rows():
        for cell in row:
            new_cell = target_sheet.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = cell.number_format
                new_cell.alignment = copy(cell.alignment)

    # Translate headers
    headers = [str(cell.value) for cell in source_sheet[1]]
    translated = translate_text_with_gemini(headers, src_lang, tgt_lang, model_name) # Pass list of headers
    for idx, val in enumerate(translated, 1):
        target_sheet.cell(row=1, column=idx).value = val or headers[idx-1] # Fallback to original header if translation fails

    # Translate content
    for col_idx in tqdm(range(1, source_sheet.max_column + 1), desc="Columns"):
        process_column(source_sheet, target_sheet, col_idx, src_lang, tgt_lang, model_name, batch_size)


def process_column(source, target, col_idx, src_lang, tgt_lang, model_name, batch_size):
    """Process column data with batching using Gemini API"""
    col_letter = get_column_letter(col_idx)
    batch, cells = [], []

    for row_idx in range(2, source.max_row + 1):
        cell = source.cell(row_idx, col_idx)
        if cell.value:
            batch.append(str(cell.value))
            cells.append((row_idx, col_idx))

    for i in range(0, len(batch), batch_size):
        translated = translate_text_with_gemini(batch[i:i+batch_size], src_lang, tgt_lang, model_name) # Translate batch with Gemini
        for (row, col), text in zip(cells[i:i+batch_size], translated):
            target.cell(row, col).value = text or source.cell(row, col).value # Fallback if translation fails

# CLI Interface
def main():
    install_dependencies()
    parser = argparse.ArgumentParser(description="Excel Translation Toolkit (Gemini API)",
                                    epilog="Example: excel_translator_gemini.py translate input.xlsx --target-lang en --model gemini-1.5-flash")
    subparsers = parser.add_subparsers(dest='command')

    # Translate command
    translate_parser = subparsers.add_parser('translate', help='Translate Excel file')
    translate_parser.add_argument('input_file', help='Input Excel file path')
    translate_parser.add_argument('--source-lang', default='auto', help='Source language code (default: auto)')
    translate_parser.add_argument('--target-lang', required=True, help='Target language code (e.g. en, ja)')
    translate_parser.add_argument('--model', default='gemini-1.5-flash',
                                help='Gemini model name (default: gemini-1.5-flash)')
    translate_parser.add_argument('--batch-size', type=int, default=DEFAULT_BATCH_SIZE,
                                help=f'Translation batch size (default: {DEFAULT_BATCH_SIZE})')
    translate_parser.add_argument('--no-post-process', action='store_false', dest='post_process',
                                help='Disable format preservation')

    # Post-process command (No changes needed in definition, but functionality is now different if called directly)
    pp_parser = subparsers.add_parser('postprocess', help='Post-process existing translation')
    pp_parser.add_argument('original_file', help='Original Excel file')
    pp_parser.add_argument('translated_file', help='Translated Excel file')

    args = parser.parse_args()

    if args.command == 'translate':
        if not os.getenv('GOOGLE_API_KEY'):
            print(f"{Fore.RED}Error: GOOGLE_API_KEY environment variable is not set.{Style.RESET_ALL}")
            print(f"{Fore.YELLOW}Please set your Gemini API key as an environment variable.{Style.RESET_ALL}")
            exit(1)

        success = translate_excel(
            args.input_file,
            args.source_lang,
            args.target_lang,
            args.model,
            args.batch_size,
            args.post_process
        )
        exit(0 if success else 1)

    elif args.command == 'postprocess': # Post-process now directly calls the improved function
        input_file_ext = os.path.splitext(args.original_file)[1] # Get extension for postprocess command as well
        success = preserve_formatting(args.original_file, args.translated_file, input_file_ext)
        exit(0 if success else 1)

    else:
        parser.print_help()
        exit(1)

if __name__ == "__main__":
    main()